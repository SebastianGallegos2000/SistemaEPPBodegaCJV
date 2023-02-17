from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils import *
def crearExcel():
  #wb = Workbook() # se crea un excel
  #wb.save('nuevo_excel.xlsx') # se crea el excel y se le coloca un nombre
  lwb = load_workbook('nuevo_excel.xlsx') # se carga el archivo excel
  ws = lwb['Epps'] # se crea la hoja en la que se va a trabajar
  ##ws.title = 'Epps' # se le cambia el nombre a la hoja
  ws['A1'] = "Bodega" # se asigna nombre a la celda A1
  ws['B1'] = "Codigo" # se asgna nombre a la celda B1
  ws['C1'] = "Descripcion" # se asgna nombre a la celda C1
  ws['D1'] = "Precio por epp" # se asgna nombre a la celda D1
  ws['E1'] = "Cantidad de epp" # se asgna nombre a la celda E1
  ws['F1'] = "Precio total" # se asgna nombre a la celda F1
  colBod = ws['A'] # se le asigna una columna entera a la variable
  colCod = ws['B'] # se le asigna una columna entera a la variable
  colDes = ws['C'] # se le asigna una columna entera a la variable
  colPre = ws['D'] # se le asigna una columna entera a la variable
  colCan = ws['E'] # se le asigna una columna entera a la variable
  colPreTotal = ws['F'] # se le asigna una columna entera a la variable
  cell_range_Bod = ws['A2':'A40'] # se le asigna un rango a una variable
  cell_range_Cod = ws['B2':'B40'] # se le asigna un rango a una variable
  cell_range_Des = ws['C2':'C40'] # se le asigna un rango a una variable
  cell_range_Pre = ws['D2':'D40'] # se le asigna un rango a una variable
  cell_range_Can = ws['E2':'E40'] # se le asigna un rango a una variable
  cell_range_PreTotal = ws['F2':'F40'] # se le asigna un rango a una variable
  print(colBod,colCod,colDes,colPre,colCan,colPreTotal) # se muestra por pantalla las columnas
  print(cell_range_Pre)

  lwb.save('nuevo_excel.xlsx') # se guardan los cambios del excel
crearExcel()